#
# SPDX-FileCopyrightText: Copyright © 2023 Idiap Research Institute <contact@idiap.ch>
#
# SPDX-FileContributor: Théophile Gentilhomme <theophile.gentilhomme@idiap.ch>
#
# SPDX-License-Identifier: GPL-3.0-only
#
# anonymization: Text ner and pii
#

import openpyxl
import pandas as pd

from anonymization.pii import Anonymize
from anonymization.pii import PIIDetection
from anonymization.pii.anonym_ops import gen_operators


class Anonymizer:
    def __init__(self, config):
        """
        Anonymizer.
        """
        self.config = config
        self.detector = PIIDetection(config)
        self.anonym = Anonymize(operators=gen_operators(self.config))

    def __call__(self, cell_value):
        """
        Perform anonymization on a cell value using the provided anonymization function.

        Parameters:
            cell_value: The value of a cell in the DataFrame.

        Returns:
            The anonymized cell value.
        """
        text = str(cell_value)
        if not text.strip():
            return text
        results = self.detector.analyse(text)
        return self.anonym.anonymise(text, results)


def load_and_anonymize(file_name, config):
    """
    Load a file into a pandas DataFrame and apply anonymization to specified columns
    or the entire DataFrame.

    Parameters:
        file_name (str): The name of the file to be loaded.
        Supported formats: CSV, Excel (XLSX), or text (TXT).
        config (dict): Anonymization configuration.

    Returns:
        None: The function saves the anonymized DataFrame
        to a new file with "anonymized" appended to the original file name
              before the extension. No explicit return value.
    """
    # Detect the file format based on the extension
    file_extension = file_name.split(".")[-1].lower()
    columns = config.get("process_columns", [])

    anonymizer = Anonymizer(config)
    anonymized_file_name = file_name.replace(
        f".{file_extension}", f"_anonymized.{file_extension}"
    )
    if file_extension == "csv":
        df = pd.read_csv(file_name)
        for c in columns:
            if c < len(df.columns):
                df.iloc[:, c] = df.iloc[:, c].apply(anonymizer)
        df.to_csv(anonymized_file_name, index=False)
    elif file_extension == "xlsx":
        # Load the Excel file
        workbook = openpyxl.load_workbook(file_name)
        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate through all rows and columns of the sheet
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
            ):
                for cell in row:
                    if cell.value and cell.column in columns:
                        anonymizer.current_column_name = cell.column_letter
                        cell.value = anonymizer(cell.value)

        # Save the modified Excel file
        workbook.save(anonymized_file_name)
    elif file_extension == "txt":
        with open(file_name, "r") as file:
            text = anonymizer(file.read())
        with open(anonymized_file_name, "w") as file:
            file.write(text)
    else:
        raise ValueError(f"Unsupported file format: {file_extension}")
    print(f"Anonymized data saved to '{anonymized_file_name}'")
